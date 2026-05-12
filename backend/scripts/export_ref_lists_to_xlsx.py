"""
SAVE TO: backend/scripts/export_ref_lists_to_xlsx.py

Export every reference list (DB-backed AND static-from-code) into a single
multi-tab Excel workbook for business-user review and cleanup.

For each list, the workbook shows:
  - The full set of available values (DB rows + static-list entries)
  - PLUS any extra values found IN USE in tx_case / tx_case_service that
    are NOT in the catalogue (likely typos or missing dropdown entries)
  - USAGE_COUNT for every row, so orphaned rows are easy to spot

Run it once. Hand the resulting xlsx to your users. They mark up the ACTION /
NEW_VALUE / MERGE_INTO_ID / NOTES columns. Send it back and we generate the
cleanup migration.

Usage:
    python backend/scripts/export_ref_lists_to_xlsx.py
        → writes ref_lists_for_review_<YYYY-MM-DD>.xlsx in the current dir

    python backend/scripts/export_ref_lists_to_xlsx.py my_path.xlsx
        → writes to the path you specify

Requirements (already in your backend venv):
    psycopg / psycopg2
    openpyxl
"""

from __future__ import annotations

import os
import sys
from datetime import date

try:
    import psycopg  # psycopg3
    PSYCOPG_VERSION = 3
except ImportError:
    import psycopg2 as psycopg  # type: ignore[no-redef]
    PSYCOPG_VERSION = 2

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================================
# Static lists — MUST stay in sync with backend/main.py REF_LIST_STATIC
# ============================================================================
STATIC_LISTS: dict[str, list[str]] = {
    "source_types":      ["DIRECT", "SUB_AGENT", "MASTER_AGENT", "GROUP", "OFFICE"],
    "import_statuses":   ["OK", "UNRESOLVED", "FLAGGED", "SCRAP"],
    "course_statuses":   ["Attending", "Enrolled", "Cancelled"],
    "deferral_codes":    ["NONE", "DEFERRED", "FEE_TRANSFERRED", "FEE_WAIVED", "NO_SERVICE"],
    "system_types":      ["Trong hệ thống", "Ngoài hệ thống"],
    "institution_types": ["DIRECT", "MASTER_AGENT", "GROUP", "OUT_OF_SYSTEM", "RMIT_VN", "OTHER_VN"],
    "bonus_events": [
        "contract_signed_date",
        "visa_received_date",
        "course_start_date",
        "enrolment_date",
        "manual_hold",
    ],
    "presales_agents": [
        "NONE", "Gia Mẫn", "Hoàng Yến", "Huỳnh Anh",
        "Lê Thị Trường An", "Trúc Quỳnh (HCM)", "Trúc Quỳnh (HN)",
    ],
    "addon_codes": [
        "EXTRA_SCHOOL", "VISITOR_VISA", "STUDY_PERMIT_RENEWAL",
        "GUARDIAN_VISA_RENEWAL", "SCHOOL_TRANSFER_DET", "CAQ",
        "GUARDIAN_HOMESTAY_CHANGE", "EXCHANGE",
    ],
    "client_types": [
        "Credential Evaluation",
        "Công tác nước ngoài",
        "Du hoc (Ghi danh + visa)", "Du hoc (Ghi danh)", "Du hoc (ghi danh + visa)",
        "Du hoc he", "Du hoc hè", "Du hoc tai cho", "Du hoc tai cho (Vietnam)",
        "Du học (Ghi danh + visa)", "Du học (Ghi danh)", "Du học (Ghi danh+visa)",
        "Du học (Nộp đơn hỗ trợ tài chính)", "Du học (chuyển trường)",
        "Du học (ghi danh + visa)", "Du học (ghi danh chuyển trường)",
        "Du học (ghi danh)", "Du học (hướng dẫn phỏng vấn)", "Du học (visa)",
        "Du học (điền đơn chuyển trường)", "Du học (điền đơn ghi danh)",
        "Du học (điền đơn visa)", "Du học (điền đơn xin học bổng)",
        "Du học (điền đơn, đăng ký lịch phỏng vấn, đóng phí SEVIS)",
        "Du học he", "Du học hè", "Du học tại chỗ", "Du học tại chỗ (VN)",
        "Du học tại chỗ (Vietnam)", "Du học/tham quan ngắn hạn theo đoàn",
        "Du lịch", "Giám hộ ở nước ngoài", "Kết hôn",
        "Người phụ thuộc ở nước ngoài", "Thay đổi Giám Hộ / Chỗ ở",
        "Thị thực tạm trú cho sinh viên sau tốt nghiệp",
        "Thị thực tạm trú cho sinh viên tốt nghiệp",
        "Travel Exemption",
        "Visa Dinh cu", "Visa Du Lịch", "Visa Du hoc only", "Visa Du học only",
        "Visa Du lịch", "Visa Giam ho", "Visa Giám hộ", "Visa Phu thuoc",
        "Visa Phụ thuộc", "Visa du học only", "Visa du lich", "Visa giám hộ",
        "Visa phụ thuộc", "Visa Định cư", "Visa định cư",
        "Điền đơn xin visa",
    ],
}


# ============================================================================
# Where each static list is used in the data, so we can count usage and
# surface in-use-only values. Format: list_name → list of (table, column).
# ============================================================================
STATIC_LIST_USAGE: dict[str, list[tuple[str, str]]] = {
    "client_types":      [("tx_case", "client_type_code")],
    "course_statuses":   [("tx_case", "course_status")],
    "deferral_codes":    [("tx_case", "deferral_code")],
    "system_types":      [("tx_case", "system_type")],
    "institution_types": [("tx_case", "institution_type")],
    "bonus_events":      [("tx_case_service", "bonus_event")],
    "source_types":      [("tx_case", "referring_source_type")],
    "import_statuses":   [("tx_case", "import_status")],
    # presales_agents: handled specially (via JOIN ref_staff)
    # addon_codes: legacy, no current usage column
}


# ============================================================================
# DB-backed lists — table, primary key, ordering, and FK columns that
# reference this table's id, used to compute USAGE_COUNT per row.
# ============================================================================
DB_LISTS: list[dict] = [
    {
        "sheet": "Countries",
        "table": "dim_country",
        "order_by": "name",
        "fk_columns": [("tx_case", "country_id")],
    },
    {
        "sheet": "Offices",
        "table": "dim_office",
        "order_by": "code",
        "fk_columns": [
            ("tx_case", "case_office_id"),
            ("tx_case", "referring_office_id"),
        ],
    },
    {
        "sheet": "Roles",
        "table": "dim_role",
        "order_by": "code",
        "fk_columns": [
            ("tx_case", "counsellor_role_id"),
            ("tx_case", "case_officer_role_id"),
        ],
    },
    {
        "sheet": "Status Split",
        "table": "ref_status_split",
        "order_by": "status",
        # tx_case.application_status is text not FK — handled separately
        "fk_columns": [],
        "text_match_column": ("tx_case", "application_status", "status"),
        # (where, what, against_local_col)
    },
    {
        "sheet": "Institutions",
        "table": "ref_institution",
        "order_by": "canonical_name",
        "fk_columns": [("tx_case", "institution_id")],
    },
    {
        "sheet": "Sub-Agents",
        "table": "ref_sub_agent",
        "order_by": "canonical_name",
        "fk_columns": [("tx_case", "referring_sub_agent_id")],
    },
    {
        "sheet": "Partners",
        "table": "ref_partner",
        "order_by": "name",
        "fk_columns": [("tx_case", "referring_partner_id")],
    },
    {
        "sheet": "Staff",
        "table": "ref_staff",
        "order_by": "canonical_name",
        "fk_columns": [
            ("tx_case", "counsellor_staff_id"),
            ("tx_case", "case_officer_staff_id"),
            ("tx_case", "presales_staff_id"),
            ("tx_case", "pre_sales_staff_id"),
            ("tx_case", "vp_staff_id"),
            ("tx_case", "target_owner_staff_id"),
        ],
    },
    {
        "sheet": "Service Fees",
        "table": "ref_service_fee",
        "order_by": "category, service_code",
        "fk_columns": [
            ("tx_case", "service_fee_id"),
            ("tx_case", "package_fee_id"),
            ("tx_case_service", "service_fee_id"),
        ],
    },
]


# Columns hidden from the cleanup workbook to reduce noise
HIDE_COLUMNS = {"created_at", "updated_at"}

# Columns appended to every sheet for business markup
MARKUP_HEADERS = ["ACTION", "NEW_VALUE", "MERGE_INTO_ID", "NOTES"]


# ============================================================================
# Styling
# ============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ACTION_FILL = PatternFill("solid", fgColor="FFF2CC")    # pale yellow — for markup
IN_USE_FILL = PatternFill("solid", fgColor="F8CBAD")    # pale orange — IN_USE_ONLY rows
ZERO_USE_FILL = PatternFill("solid", fgColor="E2EFDA")  # pale green — usage 0 (deletable?)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header_row(ws: Worksheet, last_col: int) -> None:
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def auto_size_columns(ws: Worksheet, last_col: int) -> None:
    """Approximate auto-fit, capped at 60 chars."""
    for col_idx in range(1, last_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            if cell.value is None:
                continue
            try:
                ln = len(str(cell.value))
            except Exception:
                ln = 8
            max_len = max(max_len, ln)
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def append_markup_columns(ws: Worksheet, data_col_count: int, row_count: int) -> None:
    start = data_col_count + 1
    for offset, header in enumerate(MARKUP_HEADERS):
        col_idx = start + offset
        # Header
        h = ws.cell(row=1, column=col_idx, value=header)
        h.fill = HEADER_FILL
        h.font = HEADER_FONT
        h.alignment = Alignment(horizontal="left", vertical="center")
        h.border = THIN_BORDER
        # Body — tint yellow
        for r in range(2, row_count + 2):
            c = ws.cell(row=r, column=col_idx)
            c.fill = ACTION_FILL
            c.border = THIN_BORDER


# ============================================================================
# DB helpers — bulk usage queries
# ============================================================================

def fk_count_query(table: str, column: str) -> str:
    """Group-by query: how many rows in `table` reference each value of `column`."""
    return (
        f"SELECT {column} AS k, COUNT(*) AS n "
        f"FROM {table} WHERE {column} IS NOT NULL GROUP BY {column}"
    )


def text_count_query(table: str, column: str) -> str:
    """For text-valued (non-FK) columns. Same shape, but the key is a string."""
    return fk_count_query(table, column)


def get_fk_counts_for_table(cur, fk_columns: list[tuple[str, str]]) -> dict[int, int]:
    """Sum usage counts across multiple FK columns into {target_id: total}."""
    counts: dict[int, int] = {}
    for table, column in fk_columns:
        cur.execute(fk_count_query(table, column))
        for row in cur.fetchall():
            k = row[0]
            n = row[1]
            if k is None:
                continue
            counts[k] = counts.get(k, 0) + n
    return counts


def get_text_counts(cur, table: str, column: str) -> dict[str, int]:
    """For text-valued columns. Returns {value: count}."""
    counts: dict[str, int] = {}
    cur.execute(text_count_query(table, column))
    for row in cur.fetchall():
        k = row[0]
        n = row[1]
        if k is None:
            continue
        counts[str(k)] = counts.get(str(k), 0) + n
    return counts


def get_presales_counts(cur) -> dict[str, int]:
    """Special-case: count presales usage by canonical_name (joined via FK)."""
    cur.execute(
        """
        SELECT s.canonical_name AS k, COUNT(*) AS n
          FROM tx_case c
          JOIN ref_staff s ON c.pre_sales_staff_id = s.id
         GROUP BY s.canonical_name
        """
    )
    out: dict[str, int] = {}
    for row in cur.fetchall():
        k = row[0]
        n = row[1]
        if k is None:
            continue
        out[str(k)] = n
    return out


# ============================================================================
# Sheet builders
# ============================================================================

def add_db_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[tuple],
    id_col_index: int | None,
    usage_counts: dict[int, int],
    extra_text_match: dict[str, int] | None = None,
    text_match_col_index: int | None = None,
) -> None:
    """Render a DB-backed sheet with USAGE_COUNT and (if text-matched) extra
    IN_USE_ONLY rows for values seen in tx_case but absent from the catalog."""

    keep_idx = [i for i, c in enumerate(columns) if c not in HIDE_COLUMNS]
    keep_cols = [columns[i] for i in keep_idx]

    ws = wb.create_sheet(title=sheet_name[:31])

    # Headers
    full_headers = list(keep_cols) + ["USAGE_COUNT"]
    if extra_text_match is not None:
        full_headers.append("SOURCE")
    for i, col in enumerate(full_headers, start=1):
        ws.cell(row=1, column=i, value=col)

    data_col_count = len(full_headers)
    matched_text_values: set[str] = set()

    # DB rows
    for r, row in enumerate(rows, start=2):
        for i, src_idx in enumerate(keep_idx, start=1):
            val = row[src_idx]
            if val is not None and not isinstance(val, (str, int, float, bool)):
                val = str(val)
            cell = ws.cell(row=r, column=i, value=val)
            cell.border = THIN_BORDER

        # USAGE_COUNT
        usage = 0
        if id_col_index is not None:
            row_id = row[id_col_index]
            if isinstance(row_id, int):
                usage = usage_counts.get(row_id, 0)
        # Add text-match counts too, if applicable
        if extra_text_match is not None and text_match_col_index is not None:
            text_val = row[text_match_col_index]
            if text_val is not None:
                matched_text_values.add(str(text_val))
                usage += extra_text_match.get(str(text_val), 0)

        usage_cell = ws.cell(row=r, column=len(keep_cols) + 1, value=usage)
        usage_cell.border = THIN_BORDER
        if usage == 0:
            usage_cell.fill = ZERO_USE_FILL

        if extra_text_match is not None:
            src_cell = ws.cell(row=r, column=len(keep_cols) + 2, value="DB")
            src_cell.border = THIN_BORDER

    last_data_row = len(rows) + 1

    # IN_USE_ONLY rows for text-matched lists (Status Split)
    if extra_text_match is not None and text_match_col_index is not None:
        in_use_only = [
            (v, n) for v, n in extra_text_match.items()
            if v not in matched_text_values
        ]
        in_use_only.sort(key=lambda x: -x[1])
        for v, n in in_use_only:
            last_data_row += 1
            ws.cell(row=last_data_row, column=text_match_col_index + 1, value=v)  # +1 for 1-based
            ws.cell(row=last_data_row, column=len(keep_cols) + 1, value=n)
            ws.cell(row=last_data_row, column=len(keep_cols) + 2, value="IN_USE_ONLY")
            for c_idx in range(1, len(full_headers) + 1):
                cell = ws.cell(row=last_data_row, column=c_idx)
                cell.border = THIN_BORDER
                if cell.fill == PatternFill():  # no fill yet
                    cell.fill = IN_USE_FILL

    style_header_row(ws, data_col_count)
    append_markup_columns(ws, data_col_count, last_data_row - 1)
    auto_size_columns(ws, data_col_count + len(MARKUP_HEADERS))


def add_static_sheet(
    wb: Workbook,
    sheet_name: str,
    static_values: list[str],
    usage: dict[str, int],
) -> int:
    """Render a static-list sheet. Returns total row count (incl IN_USE_ONLY)."""
    ws = wb.create_sheet(title=sheet_name[:31])

    headers = ["VALUE", "SOURCE", "USAGE_COUNT"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)

    static_set = set(static_values)
    in_use_set = set(usage.keys())
    all_values = sorted(static_set | in_use_set)

    row_idx = 1
    for v in all_values:
        row_idx += 1
        if v in static_set and v in in_use_set:
            source = "BOTH"
        elif v in static_set:
            source = "STATIC"
        else:
            source = "IN_USE_ONLY"
        n = usage.get(v, 0)
        ws.cell(row=row_idx, column=1, value=v).border = THIN_BORDER
        src_cell = ws.cell(row=row_idx, column=2, value=source)
        src_cell.border = THIN_BORDER
        n_cell = ws.cell(row=row_idx, column=3, value=n)
        n_cell.border = THIN_BORDER

        if source == "IN_USE_ONLY":
            ws.cell(row=row_idx, column=1).fill = IN_USE_FILL
            src_cell.fill = IN_USE_FILL
            n_cell.fill = IN_USE_FILL
        elif n == 0:
            n_cell.fill = ZERO_USE_FILL

    body_rows = row_idx - 1
    style_header_row(ws, 3)
    append_markup_columns(ws, 3, body_rows)
    auto_size_columns(ws, 3 + len(MARKUP_HEADERS))
    return body_rows


def add_summary_sheet(wb: Workbook, summary_rows: list[tuple]) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    headers = ["List", "Source", "Catalogue rows", "In-use only", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    for r, row in enumerate(summary_rows, start=2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN_BORDER

    style_header_row(ws, len(headers))
    auto_size_columns(ws, len(headers))

    intro_row = len(summary_rows) + 4
    intro = [
        "How to use this workbook:",
        "",
        "Each tab is one reference list used by the bonus engine.",
        "",
        "Cell colour legend:",
        "    YELLOW columns   = markup columns for your edits "
        "(ACTION, NEW_VALUE, MERGE_INTO_ID, NOTES)",
        "    ORANGE rows      = IN_USE_ONLY — the value appears in actual "
        "case data but is NOT in the master catalogue (likely typo, or a "
        "value that needs adding)",
        "    GREEN cell       = USAGE_COUNT is 0 (not referenced anywhere "
        "→ safe candidate for DELETE)",
        "",
        "ACTION values to use in the yellow ACTION column:",
        "    KEEP   = leave as-is (default — no need to write anything)",
        "    RENAME = change the value/label; put the new text in NEW_VALUE",
        "    DELETE = remove this row from the catalogue",
        "    MERGE  = redirect every reference of this row to another row's id; "
        "put the target's id in MERGE_INTO_ID",
        "    NEW    = add a new entry to the catalogue (only useful when you "
        "want to add a row that isn't already shown as IN_USE_ONLY)",
        "",
        "The 'Source' column on this Summary tab tells you whether a list "
        "lives in the database (DB table) or in the application code "
        "(STATIC). Static-list edits will require a code change to apply.",
        "",
        "Return this workbook when done. We'll generate a cleanup script that "
        "applies your edits (with a dry-run preview first).",
    ]
    for i, line in enumerate(intro):
        cell = ws.cell(row=intro_row + i, column=1, value=line)
        if i == 0:
            cell.font = Font(bold=True, size=12)
        ws.merge_cells(start_row=intro_row + i, start_column=1, end_row=intro_row + i, end_column=5)


# ============================================================================
# Main
# ============================================================================

def get_database_url() -> str:
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit(
            "DATABASE_URL environment variable is not set.\n\n"
            "Set it the same way the FastAPI backend reads it. Example for "
            "Windows PowerShell:\n"
            '    $env:DATABASE_URL="postgres://...your...railway...url..."\n'
            "Then re-run this script in the same terminal.\n"
        )
    return url


def column_index_for(columns: list[str], target: str) -> int | None:
    try:
        return columns.index(target)
    except ValueError:
        return None


def main() -> None:
    output_path = (
        sys.argv[1] if len(sys.argv) > 1
        else f"ref_lists_for_review_{date.today().isoformat()}.xlsx"
    )

    url = get_database_url()
    summary_rows: list[tuple] = []

    wb = Workbook()
    wb.remove(wb.active)

    print(f"Connecting to database (psycopg v{PSYCOPG_VERSION})…")
    with psycopg.connect(url) as conn:
        cur = conn.cursor()

        # --- DB-backed lists ----------------------------------------------
        for spec in DB_LISTS:
            sheet = spec["sheet"]
            table = spec["table"]
            order_by = spec["order_by"]
            print(f"  Exporting DB list: {sheet} ({table})")

            cur.execute(f"SELECT * FROM {table} ORDER BY {order_by}")
            if PSYCOPG_VERSION == 3:
                columns = [c.name for c in cur.description]
            else:
                columns = [c[0] for c in cur.description]
            rows = cur.fetchall()

            usage_counts = get_fk_counts_for_table(cur, spec["fk_columns"])

            extra_text: dict[str, int] | None = None
            text_match_idx: int | None = None
            in_use_only_count = 0
            if "text_match_column" in spec:
                t_table, t_col, local_col = spec["text_match_column"]
                extra_text = get_text_counts(cur, t_table, t_col)
                text_match_idx = column_index_for(columns, local_col)
                # Count of IN_USE_ONLY (values in text data not in catalog)
                if text_match_idx is not None:
                    catalog_values = {
                        str(row[text_match_idx]) for row in rows
                        if row[text_match_idx] is not None
                    }
                    in_use_only_count = sum(
                        1 for k in extra_text if k not in catalog_values
                    )

            id_idx = column_index_for(columns, "id")
            add_db_sheet(
                wb, sheet, columns, rows,
                id_col_index=id_idx,
                usage_counts=usage_counts,
                extra_text_match=extra_text,
                text_match_col_index=text_match_idx,
            )
            summary_rows.append((
                sheet, "DB table",
                len(rows),
                in_use_only_count,
                "",
            ))

        # --- Static lists -------------------------------------------------
        for name, values in STATIC_LISTS.items():
            print(f"  Exporting static list: {name}")
            if name == "presales_agents":
                usage = get_presales_counts(cur)
            elif name in STATIC_LIST_USAGE:
                usage = {}
                for table, column in STATIC_LIST_USAGE[name]:
                    for k, n in get_text_counts(cur, table, column).items():
                        usage[k] = usage.get(k, 0) + n
            else:
                usage = {}

            static_set = set(values)
            in_use_only_count = sum(1 for k in usage if k not in static_set)

            add_static_sheet(wb, name, values, usage)
            summary_rows.append((
                name, "STATIC (main.py)",
                len(values),
                in_use_only_count,
                "Code change needed to apply edits",
            ))

        cur.close()

    add_summary_sheet(wb, summary_rows)

    wb.save(output_path)
    print(f"\nDone. Wrote {output_path}")
    print(f"Tabs: {len(summary_rows) + 1} (1 Summary + {len(summary_rows)} list tabs)")


if __name__ == "__main__":
    main()
