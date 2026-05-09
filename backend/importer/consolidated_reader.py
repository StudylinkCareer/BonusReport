"""
backend/importer/consolidated_reader.py

Reader for the going-forward consolidated CRM file format.
(Initially: Report_Jul_2023_-_Apr_2026_1.xlsx; format will continue.)

Differences from the per-month reader.py:
  * Header at row 1 (per-month: row 2 or 3)
  * 16 columns, no 'No.' column, no section dividers
  * Date strings in d/m/yyyy or dd/mm/yyyy format are parsed at read time
    so the transformer always sees datetime objects (not strings)
  * Run period is derived per row by the orchestrator, NOT from the
    filename
  * Filtering at read time: pass any combination of staff/date criteria
    and only matching rows are yielded

The reader is pure I/O — no DB calls. Filter sets are built upstream by
the orchestrator (which has cursor access for role/office lookups).

Public API
----------
iter_filtered_rows(path, *, staff_names_lower=None, contract_signed_from=None,
                   contract_signed_to=None, visa_received_from=None,
                   visa_received_to=None, course_start_from=None,
                   course_start_to=None, limit=None) -> Iterator[RawRow]

The yielded RawRow is the same dataclass used by reader.py — a drop-in
input for transformer.transform_row().
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Optional

from openpyxl import load_workbook

# Reuse the existing RawRow so transformer.py doesn't need to know which
# reader produced the row.
from backend.importer.reader import RawRow


# ---------------------------------------------------------------------------
# Sheet expectations
# ---------------------------------------------------------------------------

EXPECTED_SHEET_NAME = "Student Contract"

# Columns we recognise. Anything else is ignored. Dates among these get
# parsed to datetime in the yielded RawRow so the transformer doesn't have
# to deal with mixed types.
DATE_COLUMNS = ("Contract Signed Date", "Visa Received Date", "Course Start Date")

# Fields the reader insists on seeing in the header row. Anything in the
# 16-col format must include these (the rest are optional/nullable).
REQUIRED_COLUMNS = (
    "Student Name", "Contract ID", "Application Report Status",
    "Counsellor Name", "Case Officer Name",
)


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------

class UnexpectedSheetError(ValueError):
    """The expected 'Student Contract' sheet was not found."""


class UnexpectedHeaderError(ValueError):
    """The header row is missing one or more required columns."""


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

# d/m/yyyy or dd/mm/yyyy — Vietnamese convention. NBSP-tolerant via
# upstream replacement, no year zero-pad needed (always 4 digits).
_DMY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _parse_date_cell(v: Any) -> Optional[datetime]:
    """Normalise a cell value to a datetime (or None).

    openpyxl returns datetime/date for Excel-typed cells, str for text-typed
    cells. Strings come in d/m/yyyy or dd/mm/yyyy. Returns None for blank
    or unparseable input.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        # date subclass returns False for isinstance(_, datetime), so this
        # branch only catches plain dates
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.replace("\u00a0", " ").strip()
        if not s:
            return None
        m = _DMY_RE.match(s)
        if not m:
            return None
        d_, mo_, y_ = m.groups()
        try:
            return datetime(int(y_), int(mo_), int(d_))
        except ValueError:
            return None
    # Anything else (int, float, etc.) is invalid for a date column
    return None


# ---------------------------------------------------------------------------
# Filter helpers
# ---------------------------------------------------------------------------

def _date_in_range(
    value: Optional[datetime],
    lo: Optional[datetime],
    hi: Optional[datetime],
) -> bool:
    """Return True if value passes [lo, hi] bounds (inclusive on both).

    Semantics:
      - If neither bound supplied: any value passes (including None).
      - If either bound is supplied: None values are EXCLUDED (the user
        clearly cared about this date column, so empty doesn't qualify).
      - lo only: value >= lo.
      - hi only: value <= hi.
      - both: lo <= value <= hi.
    """
    if lo is None and hi is None:
        return True
    if value is None:
        return False
    if lo is not None and value < lo:
        return False
    if hi is not None and value > hi:
        return False
    return True


def _name_matches(value: Any, allowed_lower: set[str]) -> bool:
    """Case-insensitive, NBSP-tolerant name match against an allowed set.

    The set holds lowercased names (canonical + any aliases). Empty/None
    cell values never match.
    """
    if not value:
        return False
    s = str(value).replace("\u00a0", " ").strip().lower()
    return bool(s) and s in allowed_lower


# ---------------------------------------------------------------------------
# Header location
# ---------------------------------------------------------------------------

def _read_header(ws) -> dict[str, int]:
    """Build {header_text: column_index_1based} from row 1.

    Raises UnexpectedHeaderError if any required column is missing.
    """
    header_map: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v is None:
            continue
        text = str(v).strip()
        if text:
            header_map[text] = col_idx
    missing = [c for c in REQUIRED_COLUMNS if c not in header_map]
    if missing:
        raise UnexpectedHeaderError(
            f"Header row 1 is missing required columns: {missing}. "
            f"Saw: {sorted(header_map.keys())}"
        )
    return header_map


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def iter_filtered_rows(
    path: Path | str,
    *,
    staff_names_lower: Optional[set[str]] = None,
    contract_signed_from: Optional[datetime] = None,
    contract_signed_to: Optional[datetime] = None,
    visa_received_from: Optional[datetime] = None,
    visa_received_to: Optional[datetime] = None,
    course_start_from: Optional[datetime] = None,
    course_start_to: Optional[datetime] = None,
    limit: Optional[int] = None,
) -> Iterator[RawRow]:
    """Yield each data row from the consolidated CRM file, applying filters.

    Args
    ----
    path : str or Path to the .xlsx file.
    staff_names_lower : optional set of lowercased name strings; row passes
        only if its Counsellor Name OR Case Officer Name (lowercased) is
        in this set. None means no staff filter.
    *_from / *_to : inclusive date bounds for the three dated columns.
        See _date_in_range docstring for None semantics.
    limit : if given, stop yielding after this many rows have passed all
        filters. Useful for sampling.

    Yields
    ------
    RawRow with:
        row_number    — 1-based spreadsheet row number (diagnostics only)
        section_label — always None (no section dividers in this format)
        data          — {header_text: cell_value} for every header column;
                        date columns parsed to datetime, others raw

    Raises
    ------
    UnexpectedSheetError, UnexpectedHeaderError, FileNotFoundError.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    yielded = 0
    try:
        if EXPECTED_SHEET_NAME not in wb.sheetnames:
            raise UnexpectedSheetError(
                f"Expected sheet {EXPECTED_SHEET_NAME!r} in {path.name}, "
                f"found: {wb.sheetnames}"
            )
        ws = wb[EXPECTED_SHEET_NAME]
        header_map = _read_header(ws)

        for row_idx in range(2, ws.max_row + 1):
            # Build the data dict
            data: dict[str, Any] = {}
            for header, col in header_map.items():
                v = ws.cell(row=row_idx, column=col).value
                if header in DATE_COLUMNS:
                    v = _parse_date_cell(v)
                data[header] = v

            # Skip blank rows (no contract id = nothing to import)
            if not data.get("Contract ID"):
                continue

            # Staff filter
            if staff_names_lower is not None:
                if not (_name_matches(data.get("Counsellor Name"), staff_names_lower)
                        or _name_matches(data.get("Case Officer Name"), staff_names_lower)):
                    continue

            # Date filters
            if not _date_in_range(data.get("Contract Signed Date"),
                                  contract_signed_from, contract_signed_to):
                continue
            if not _date_in_range(data.get("Visa Received Date"),
                                  visa_received_from, visa_received_to):
                continue
            if not _date_in_range(data.get("Course Start Date"),
                                  course_start_from, course_start_to):
                continue

            yield RawRow(
                row_number=row_idx,
                section_label=None,
                data=data,
            )
            yielded += 1
            if limit is not None and yielded >= limit:
                return
    finally:
        wb.close()
