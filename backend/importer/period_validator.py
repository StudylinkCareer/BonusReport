"""
backend/importer/period_validator.py

Reads the period (year, month) from inside a CRM report and validates it
against an externally-supplied period (typically derived from the filename).

Two report formats are supported, both auto-detected:

  Format 1 — Vietnamese-titled (Báo_cáo_*.xlsx):
    Cell A1: "Em gửi báo cáo hồ sơ đóng tháng MM/YYYY của <staff> như sau:"
    Pattern: 'tháng' followed by M/YYYY or MM/YYYY.

  Format 2 — English-titled (*report_of_closed_file_*.xlsx):
    Cell B1: "<staff>'s report of closed file in <Month> <YYYY>"
    Pattern: 'in <MonthName> <YYYY>'.

The scan covers the first 3 rows × first 5 columns of the first sheet
to absorb minor layout drift.

Robustness:
  - NFC-normalises Unicode so 'á' (precomposed) and 'a' + combining-acute
    both match.
  - Replaces non-breaking spaces (U+00A0) with regular spaces — these
    appear in some reports between words and would otherwise break the
    regex.
  - Case-insensitive English month names.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl


ENGLISH_MONTH_NAMES = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
}

# Vietnamese: "tháng 11/2024" or "tháng 1/2024" (also tolerates a hyphen)
_VN_PATTERN = re.compile(r'th[áa]ng\s*(\d{1,2})[/\-](\d{4})', re.IGNORECASE)

# English: "in March 2025" or "in December 2024"
_EN_PATTERN = re.compile(
    r'in\s+(' + '|'.join(ENGLISH_MONTH_NAMES.keys()) + r')\s+(\d{4})',
    re.IGNORECASE,
)


class ReportPeriodNotFoundError(ValueError):
    """No recognisable period header in the file."""


class ReportPeriodMismatchError(ValueError):
    """File's embedded period disagrees with the externally-supplied one."""


@dataclass(frozen=True)
class ReportPeriod:
    year: int
    month: int

    def __str__(self) -> str:
        return f"{self.year}-{self.month:02d}"


def _normalize(text: str) -> str:
    """NFC-normalise + replace non-breaking spaces so regexes match cleanly."""
    text = unicodedata.normalize('NFC', text)
    text = text.replace('\u00a0', ' ')
    return text


def _scan_text_for_period(text: str) -> ReportPeriod | None:
    """Try Vietnamese then English patterns. Return None if neither matches."""
    text = _normalize(text)

    m = _VN_PATTERN.search(text)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if 1 <= month <= 12 and 2020 <= year <= 2099:
            return ReportPeriod(year=year, month=month)

    m = _EN_PATTERN.search(text)
    if m:
        month = ENGLISH_MONTH_NAMES[m.group(1).lower()]
        year = int(m.group(2))
        if 2020 <= year <= 2099:
            return ReportPeriod(year=year, month=month)

    return None


def parse_report_period(path: Path | str) -> ReportPeriod:
    """
    Open the xlsx and find the embedded period header.

    Scans the top-left 3 rows × 5 columns of the first sheet.

    Raises:
        ReportPeriodNotFoundError — no recognisable header found.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for cell in (row or ())[:5]:
                if not isinstance(cell, str):
                    continue
                period = _scan_text_for_period(cell)
                if period is not None:
                    return period
    finally:
        wb.close()

    raise ReportPeriodNotFoundError(
        f"Could not find a period header in {path.name}. Expected one of:\n"
        f"  • 'tháng M/YYYY' (Vietnamese format, usually cell A1)\n"
        f"  • 'in <Month> YYYY' (English format, usually cell B1)\n"
        f"in the first 3 rows of the first sheet."
    )


def validate_report_period(
    path: Path | str,
    *,
    expected_year: int,
    expected_month: int,
) -> ReportPeriod:
    """
    Confirm the report's embedded period matches an externally-supplied one.

    Returns the parsed period on success.

    Raises:
        ReportPeriodNotFoundError — header missing in the file.
        ReportPeriodMismatchError — header found but doesn't match.
    """
    period = parse_report_period(path)
    if period.year != expected_year or period.month != expected_month:
        raise ReportPeriodMismatchError(
            f"Period mismatch in {Path(path).name}: "
            f"filename says {expected_year}-{expected_month:02d}, "
            f"but the report header inside the file says {period}."
        )
    return period
