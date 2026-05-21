"""
backend/importer/reader.py

Walk a 'report of closed file' xlsx and yield each data row as a RawRow.

Reader v2 changes from previous version:
  * Header column names are resolved through ref_column_alias so source
    files with SBS table-prefix variants (e.g. 'Student File::Refer
    Source Agent') work without source-side cleanup.
  * Date cells are parsed at read time. If a "date" cell contains a
    d/m/yyyy string instead of an Excel date, it's parsed into a datetime.
    Unparseable text in a date column becomes None and the row gets a
    note from the transformer.
  * Reader still has no DB awareness beyond the alias lookup — the
    orchestrator provides a cursor for that.

The reader has no knowledge of business rules. Its jobs are:
  * Open the file and validate the sheet name.
  * Locate the header row.
  * Resolve every header text through ref_column_alias to get its
    canonical name.
  * Yield each data row, skipping blank rows, header rows, and section
    divider rows. The most recent section divider is attached to each
    yielded row as metadata.

Section divider behaviour is unchanged: rows where col A has text but
col B is blank update the running section_label.
"""

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator, Optional

from openpyxl import load_workbook


log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Public types
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class RawRow:
    """One data row from a closed-file report.

    Fields
    ------
    row_number     : 1-based spreadsheet row number (for diagnostics).
    section_label  : The most recent section divider seen above this row,
                     e.g. 'Closed files', 'Enrolled'. None if no divider.
    data           : {canonical_header_name: cell_value} for every column
                     whose header had non-blank text. Unrecognised headers
                     pass through as-is. Date columns are parsed to
                     datetime if possible; unparseable strings become None.
    """
    row_number: int
    section_label: Optional[str]
    data: dict[str, Any]


@dataclass(frozen=True)
class FilenameInfo:
    """Metadata extracted from a closed-file report's filename."""
    year: int
    month: int


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------

class FilenameParseError(ValueError):
    """The filename does not contain a recognisable Month_Year suffix."""


class HeaderRowNotFoundError(ValueError):
    """No header row was found in the first few rows of the sheet."""


class UnexpectedSheetError(ValueError):
    """The active sheet is not named 'Student Contract'."""


# ---------------------------------------------------------------------------
# Filename parsing (unchanged from v1)
# ---------------------------------------------------------------------------

_MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

_FILENAME_PATTERN = re.compile(
    r"(january|february|march|april|may|june|july|august|september|october|november|december)"
    r"[^a-zA-Z0-9]*(\d{4})\.xlsx?$",
    re.IGNORECASE,
)


def parse_filename(path: Path | str) -> FilenameInfo:
    """Extract year and month from a closed-file report's filename."""
    name = Path(path).name
    match = _FILENAME_PATTERN.search(name)
    if not match:
        raise FilenameParseError(
            f"Cannot extract month/year from filename: {name!r}"
        )
    month_text, year_text = match.groups()
    return FilenameInfo(
        year=int(year_text),
        month=_MONTH_NAMES[month_text.lower()],
    )


# ---------------------------------------------------------------------------
# Column alias resolution
# ---------------------------------------------------------------------------

# Canonical column names the transformer references. These match the
# canonical_name column values in ref_column_alias.
CANONICAL_DATE_COLUMNS = (
    "Contract Signed Date",
    "Visa Received Date",
    "Course Start Date",
)


def _load_column_aliases(cursor) -> dict[str, str]:
    """Load ref_column_alias into a lookup dict.

    Returns {normalized_alias: canonical_name}, where normalized_alias
    is lower(trim(alias)). The reader normalizes its incoming headers
    the same way before looking up.
    """
    cursor.execute(
        "SELECT alias, canonical_name FROM ref_column_alias"
    )
    rows = cursor.fetchall()
    aliases: dict[str, str] = {}
    for row in rows:
        norm = row["alias"].strip().lower()
        aliases[norm] = row["canonical_name"]
    log.info("Loaded %d column aliases from ref_column_alias.", len(aliases))
    return aliases


def _resolve_header(header_text: str, aliases: dict[str, str]) -> str:
    """Resolve one header text through the alias map.

    Returns the canonical_name if a match is found, else returns the
    input text unchanged (stripped of leading/trailing whitespace).
    Case-insensitive, whitespace-tolerant match.
    """
    norm = header_text.strip().lower()
    return aliases.get(norm, header_text.strip())


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

# Vietnamese-convention d/m/yyyy or dd/mm/yyyy. Year is always 4 digits.
_DMY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _parse_date_cell(value: Any) -> Optional[datetime]:
    """Normalise a cell value to a datetime (or None).

    openpyxl returns datetime/date for Excel-typed date cells. For text-
    typed cells containing dates (Vietnamese convention d/m/yyyy or
    dd/mm/yyyy), we parse them here so the transformer always sees a
    datetime.

    Returns None for blank or unparseable input. NBSP-tolerant.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        s = value.replace("\u00a0", " ").strip()
        if not s:
            return None
        m = _DMY_RE.match(s)
        if not m:
            return None
        d_, mo_, y_ = m.groups()
        try:
            return datetime(int(y_), int(mo_), int(d_))
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Sheet structure
# ---------------------------------------------------------------------------

EXPECTED_SHEET_NAME = "Student Contract"
HEADER_SEARCH_DEPTH = 5  # rows


def _find_header_row(ws, aliases: dict[str, str]) -> tuple[int, dict[str, int]]:
    """Scan the first few rows for the header row.

    A header row is identified by 'No.' in column A and 'Student Name'
    in column B (after alias resolution).

    Returns (row_index_1based, {canonical_header_name: column_index_1based}).
    """
    for row_idx in range(1, HEADER_SEARCH_DEPTH + 1):
        a = ws.cell(row=row_idx, column=1).value
        b = ws.cell(row=row_idx, column=2).value
        if a is None or b is None:
            continue
        # Resolve A and B through aliases before comparing
        a_canon = _resolve_header(str(a), aliases)
        b_canon = _resolve_header(str(b), aliases)
        if a_canon == "No." and b_canon == "Student Name":
            header_map: dict[str, int] = {}
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                text = str(cell_value).strip()
                if not text:
                    continue
                canonical = _resolve_header(text, aliases)
                if canonical in header_map:
                    # Two headers in this file resolve to the same canonical.
                    # Warn and let the later one win (consistent with dict
                    # iteration order).
                    log.warning(
                        "Header conflict in row %d: both %r and an earlier "
                        "column resolve to %r. Using the latest occurrence.",
                        row_idx, text, canonical,
                    )
                header_map[canonical] = col_idx
            return row_idx, header_map
    raise HeaderRowNotFoundError(
        f"No header row found in first {HEADER_SEARCH_DEPTH} rows of "
        f"sheet {ws.title!r}"
    )


def _is_blank(value: Any) -> bool:
    """Treat None and whitespace-only strings as blank."""
    return value is None or (isinstance(value, str) and not value.strip())


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def iter_data_rows(cursor, path: Path | str) -> Iterator[RawRow]:
    """Yield each data row from a closed-file report.

    The caller provides a cursor so the reader can load ref_column_alias.
    The cursor is used for that one lookup only — no per-row DB calls.

    Skipped rows
    ------------
    * Rows above the header.
    * The header row itself.
    * Completely blank rows.
    * Section divider rows (col A has text, col B blank).

    Yielded rows
    ------------
    Every row where col B contains a value. The data dict is keyed by
    canonical column name (after ref_column_alias resolution). Date
    columns are parsed to datetime if possible.
    """
    path = Path(path)
    aliases = _load_column_aliases(cursor)
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        if ws.title != EXPECTED_SHEET_NAME:
            raise UnexpectedSheetError(
                f"Expected sheet {EXPECTED_SHEET_NAME!r} in {path.name}, "
                f"got {ws.title!r}"
            )

        header_row_idx, header_map = _find_header_row(ws, aliases)
        log.info(
            "Header row %d resolved to %d canonical column(s) in %s",
            header_row_idx, len(header_map), path.name,
        )
        current_section: Optional[str] = None

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            a = ws.cell(row=row_idx, column=1).value
            b = ws.cell(row=row_idx, column=2).value

            # Blank row -> skip
            if _is_blank(a) and _is_blank(b):
                continue

            # Section divider -> update label and skip
            if not _is_blank(a) and _is_blank(b):
                current_section = str(a).strip()
                continue

            # Data row -> build dict by canonical header name
            data: dict[str, Any] = {}
            for canonical_name, col_idx in header_map.items():
                raw_value = ws.cell(row=row_idx, column=col_idx).value
                if canonical_name in CANONICAL_DATE_COLUMNS:
                    data[canonical_name] = _parse_date_cell(raw_value)
                else:
                    data[canonical_name] = raw_value

            yield RawRow(
                row_number=row_idx,
                section_label=current_section,
                data=data,
            )
    finally:
        wb.close()
