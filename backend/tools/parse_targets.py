"""
parse_targets.py — Parse StudyLink monthly target files into a flat CSV.

Usage:
    python parse_targets.py <input_folder> [output_csv]

Outputs CSV with columns:
    source_file, sheet_name, staff_name_short, year, month,
    target_type, target_value, target_unit, raw_value, notes

target_type values:
    CONTRACT, ENROLMENT, CANCELLED, TELESALES, ENROL_SUMMER

target_unit values:
    COUNT      — integer target (e.g. 4 contracts/month)
    PERCENT    — percentage of contracts (e.g. 15)

Format A (monthly office targets):
    - Title row identifies operative month(s)
    - Section headers: "CHỈ TIÊU HỢP ĐỒNG", "CHỈ TIÊU HỌC SINH NHẬP HỌC", etc.
    - Office sub-headers: "VP.HCM", "VP.HN", "VP.ĐN", "VP Melbourne"
    - Staff rows with two month columns of values
    - Pair notation "(w/X)" is stripped

Format B (annual sub-agent targets):
    - Filename contains "Sub" prefix
    - Two staff sections, 12 month columns Jan-Dec
"""
from __future__ import annotations
import csv
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import Iterator
from openpyxl import load_workbook

VIETNAMESE_MONTHS = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    'thang 1': 1, 'thang 2': 2, 'thang 3': 3, 'thang 4': 4, 'thang 5': 5,
    'thang 6': 6, 'thang 7': 7, 'thang 8': 8, 'thang 9': 9, 'thang 10': 10,
    'thang 11': 11, 'thang 12': 12,
    'thang 01': 1, 'thang 02': 2, 'thang 03': 3, 'thang 04': 4, 'thang 05': 5,
    'thang 06': 6, 'thang 07': 7, 'thang 08': 8, 'thang 09': 9,
}

SECTION_PATTERNS = [
    (re.compile(r'CHỈ TIÊU TELESALES', re.IGNORECASE), 'TELESALES'),
    (re.compile(r'CHỈ TIÊU HỢP ĐỒNG', re.IGNORECASE), 'CONTRACT'),
    (re.compile(r'CHỈ TIÊU HỌC SINH.*NHẬP\s+HỌC', re.IGNORECASE), 'ENROLMENT'),
    (re.compile(r'CHỈ TIÊU CANCELLED', re.IGNORECASE), 'CANCELLED'),
    (re.compile(r'CHỈ TIÊU.*DU HỌC HÈ', re.IGNORECASE), 'ENROL_SUMMER'),
]

OFFICE_PREFIXES = ['VP.HCM', 'VP HCM', 'VP.HN', 'VP HN', 'VP. ĐN', 'VP.ĐN',
                   'VP ĐN', 'VP Melbourne', 'VP. Melbourne', 'VP.Melbourne']

TOTAL_PREFIX = 'TỔNG CHỈ TIÊU'


def fold_ascii(s: str) -> str:
    """Strip Vietnamese diacritics for case-insensitive match."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    return ''.join(c for c in s if not unicodedata.combining(c)).lower().strip()


def is_section_header(cell: str) -> str | None:
    if not isinstance(cell, str):
        return None
    for pat, code in SECTION_PATTERNS:
        if pat.search(cell):
            return code
    return None


def is_office_header(cell: str) -> bool:
    if not isinstance(cell, str):
        return False
    folded = fold_ascii(cell)
    for prefix in OFFICE_PREFIXES:
        if folded == fold_ascii(prefix):
            return True
    return False


def is_total_row(cell: str) -> bool:
    if not isinstance(cell, str):
        return False
    return cell.strip().upper().startswith('TỔNG')


def clean_staff_name(raw: str) -> str:
    """Strip pair notation '(w/X)', '(với X)', '(bao gồm ...)' from staff name."""
    if not isinstance(raw, str):
        return ''
    name = raw.strip()
    # Strip parenthetical suffixes
    name = re.sub(r'\s*\([^)]*\)\s*', ' ', name).strip()
    # Strip CO designation suffix like "(CO cho VPĐN)" — already handled by paren strip
    return name


def extract_months_from_title(title: str) -> list[tuple[int, int]]:
    """Extract (year, month) pairs from sheet title row.

    Examples:
        "BẢNG CHỈ TIÊU ... THÁNG 7 - THÁNG 8/2023" → [(2023, 7), (2023, 8)]
        "BẢNG CHỈ TIÊU ... THÁNG 5 và 6/2024"     → [(2024, 5), (2024, 6)]
        "BẢNG CHỈ TIÊU ... THÁNG 07 và 08/2025"   → [(2025, 7), (2025, 8)]
        "BẢNG CHỈ TIÊU ... THÁNG 09/2025"         → [(2025, 9)]
        "BẢNG CHỈ TIÊU ... THÁNG 4/2026"          → [(2026, 4)]
    """
    if not isinstance(title, str):
        return []
    t = title.upper()

    year_m = re.search(r'/(\d{4})', t)
    if not year_m:
        year_m = re.search(r'(20\d{2})', t)
    if not year_m:
        return []
    year = int(year_m.group(1))

    # Strip the year so it doesn't get parsed as a month number
    t_no_year = re.sub(r'/?\s*' + re.escape(year_m.group(0)), ' ', t)

    months: list[int] = []
    # Primary: any "THÁNG N" mentions
    for m in re.findall(r'TH[ÁA]NG\s+(\d{1,2})', t_no_year):
        v = int(m)
        if 1 <= v <= 12 and v not in months:
            months.append(v)
    # Secondary: trailing "và N" or "- N" patterns (multi-month titles)
    for m in re.findall(r'(?:VÀ|VA|-)\s+(\d{1,2})(?!\d)', t_no_year):
        v = int(m)
        if 1 <= v <= 12 and v not in months:
            months.append(v)

    return [(year, m) for m in sorted(months)]


def extract_column_months(header_row: list, default_year: int) -> dict[int, tuple[int, int] | None]:
    """For Format A: parse column headers like 'THÁNG 7' → (year, 7).
    Returns dict mapping column index → (year, month) or None.
    """
    out = {}
    for i, cell in enumerate(header_row):
        if not isinstance(cell, str):
            out[i] = None
            continue
        m = re.search(r'TH[ÁA]NG\s+(\d{1,2})', cell, re.IGNORECASE)
        if m:
            out[i] = (default_year, int(m.group(1)))
        else:
            out[i] = None
    return out


def parse_value(raw, target_type: str) -> tuple[float | None, str | None, str | None]:
    """Return (target_value, target_unit, notes).

    target_unit: 'COUNT' or 'PERCENT'
    For CANCELLED rows, raw is typically a percentage string.
    For numeric rows, raw is a number.
    Returns (None, None, None) if no parseable value.
    """
    if raw is None or raw == '-' or raw == '':
        return (None, None, None)

    if isinstance(raw, (int, float)):
        return (float(raw), 'COUNT', None)

    if isinstance(raw, str):
        s = raw.strip()
        if not s or s == '-':
            return (None, None, None)

        # Percentage form: "15% trên tổng số hợp đồng đã ký trong mỗi tháng"
        pct = re.search(r'(\d+(?:[.,]\d+)?)\s*%', s)
        if pct:
            pct_val = float(pct.group(1).replace(',', '.'))
            note = s if len(s) > 5 else None
            return (pct_val, 'PERCENT', note)

        # Plain number as string
        try:
            return (float(s.replace(',', '.')), 'COUNT', None)
        except ValueError:
            pass

    return (None, None, None)


def detect_format(wb_path: Path) -> str:
    """Detect 'A' (monthly office) or 'B' (annual sub-agent) by filename."""
    fname = wb_path.name.lower()
    if 'sub' in fname and ('an' in fname or 'lợi' in fname.lower() or 'loi' in fname):
        return 'B'
    return 'A'


def parse_format_a_sheet(ws, source_file: str, sheet_name: str) -> Iterator[dict]:
    """Yield target rows from a Format A sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Title row (R1) — extract months
    title = rows[0][0] if rows[0] else None
    sheet_months = extract_months_from_title(title)

    if not sheet_months:
        # Some sheets may have their month info elsewhere; skip if untitled
        return

    default_year = sheet_months[0][0]

    # Find sub-header row with month labels (typically R3) — but also fallback to title months
    # We'll use sheet_months as the operative period regardless of column header labels,
    # since we observed column labels are sometimes template residue.
    # Convention: Column index 1 → first month in sheet_months, Column index 2 → second.
    n_months = len(sheet_months)

    section = None
    for i, row in enumerate(rows):
        if not row or all(c is None for c in row):
            continue

        cell0 = row[0]
        cell0_str = str(cell0) if cell0 is not None else ''

        # Section header detection
        sec = is_section_header(cell0_str)
        if sec:
            section = sec
            continue

        # Skip office headers, totals, notes
        if is_office_header(cell0_str) or is_total_row(cell0_str):
            continue
        if cell0_str.startswith('Lưu ý') or 'GHI CHÚ' in cell0_str.upper():
            continue
        if not section:
            continue
        if not cell0_str or cell0_str.strip() in ('VĂN PHÒNG',):
            continue

        # This is a staff row
        staff_name = clean_staff_name(cell0_str)
        if not staff_name:
            continue

        notes_cell = row[3] if len(row) > 3 else None
        notes = str(notes_cell).strip() if notes_cell else None

        # Cancellation rates apply to all months covered by the sheet
        # (it's a rate, not a per-month count). Read whichever column has a value.
        if section == 'CANCELLED':
            raw1 = row[1] if len(row) > 1 else None
            raw2 = row[2] if len(row) > 2 else None
            raw = raw1 if raw1 not in (None, '', '·', '-') else raw2
            value, unit, val_notes = parse_value(raw, section)
            if value is not None:
                for yr, mo in sheet_months:
                    yield {
                        'source_file': source_file,
                        'sheet_name': sheet_name,
                        'staff_name_short': staff_name,
                        'year': yr,
                        'month': mo,
                        'target_type': section,
                        'target_value': value,
                        'target_unit': unit,
                        'raw_value': str(raw) if raw is not None else '',
                        'notes': notes or val_notes or '',
                    }
            continue

        # Numeric targets:
        #   - Single-month sheet: col 2 is operative (col 1 is a baseline ref).
        #     Fall back to col 1 if col 2 is empty (covers edge cases).
        #   - Two-month sheet: col 1 = first month, col 2 = second month.
        if len(sheet_months) == 1:
            yr, mo = sheet_months[0]
            raw2 = row[2] if len(row) > 2 else None
            raw1 = row[1] if len(row) > 1 else None
            raw = raw2 if raw2 not in (None, '', '·', '-') else raw1
            col_month_pairs = [(None, (yr, mo), raw)]
        else:
            col_month_pairs = []
            for col_idx, ym in zip([1, 2], sheet_months):
                raw = row[col_idx] if col_idx < len(row) else None
                col_month_pairs.append((col_idx, ym, raw))

        for _, (yr, mo), raw in col_month_pairs:
            value, unit, val_notes = parse_value(raw, section)
            if value is not None:
                yield {
                    'source_file': source_file,
                    'sheet_name': sheet_name,
                    'staff_name_short': staff_name,
                    'year': yr,
                    'month': mo,
                    'target_type': section,
                    'target_value': value,
                    'target_unit': unit,
                    'raw_value': str(raw) if raw is not None else '',
                    'notes': notes or val_notes or '',
                }


def parse_format_b_sheet(ws, source_file: str, sheet_name: str, year: int) -> Iterator[dict]:
    """Yield target rows from a Format B (sub-agent) sheet."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Walk: find rows where col0 is a person name and col1 is 'CHỈ TIÊU'
    # Then the next rows have month headers and target values
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        if not row:
            i += 1
            continue
        c0 = row[0]
        c1 = row[1] if len(row) > 1 else None

        # Detect staff section header
        if isinstance(c0, str) and isinstance(c1, str) and c1.strip() == 'CHỈ TIÊU':
            staff_name = c0.strip()
            # Next row should be month labels (Jan, Feb, ...)
            if i + 1 < n:
                month_row = rows[i + 1]
                month_cols = []
                for j, mc in enumerate(month_row):
                    if isinstance(mc, str):
                        key = fold_ascii(mc)
                        if key in VIETNAMESE_MONTHS:
                            month_cols.append((j, VIETNAMESE_MONTHS[key]))
                # Look for value rows below the month header
                k = i + 2
                while k < n:
                    vrow = rows[k]
                    if not vrow or vrow[0] is None:
                        k += 1
                        continue
                    label = str(vrow[0]).strip() if vrow[0] else ''
                    # Section label?
                    sec = is_section_header(label)
                    if sec:
                        # Next data row
                        k += 1
                        if k < n and rows[k]:
                            data_row = rows[k]
                            for col_idx, mo in month_cols:
                                if col_idx < len(data_row):
                                    raw = data_row[col_idx]
                                    value, unit, val_notes = parse_value(raw, sec)
                                    if value is not None:
                                        yield {
                                            'source_file': source_file,
                                            'sheet_name': sheet_name,
                                            'staff_name_short': staff_name,
                                            'year': year,
                                            'month': mo,
                                            'target_type': sec,
                                            'target_value': value,
                                            'target_unit': unit,
                                            'raw_value': str(raw) if raw is not None else '',
                                            'notes': '',
                                        }
                            k += 1
                            continue
                    # Hit another staff section header? back out
                    if isinstance(vrow[0], str) and len(vrow) > 1 and isinstance(vrow[1], str) and vrow[1].strip() == 'CHỈ TIÊU':
                        break
                    k += 1
                i = k
                continue
        i += 1


def parse_year_from_filename(fname: str) -> int | None:
    """Extract year from filename like 'StudyLink_Business_Target_2025.xlsx'."""
    m = re.search(r'(20\d{2})', fname)
    return int(m.group(1)) if m else None


def parse_file(path: Path) -> list[dict]:
    fmt = detect_format(path)
    rows = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f'  ERROR loading {path.name}: {e}', file=sys.stderr)
        return []

    if fmt == 'B':
        year = parse_year_from_filename(path.name)
        if year is None:
            print(f'  WARN: cannot extract year from {path.name}', file=sys.stderr)
            return []
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows.extend(parse_format_b_sheet(ws, path.name, sn, year))
    else:
        for sn in wb.sheetnames:
            ws = wb[sn]
            # Skip blank Sheet2/Sheet3
            if ws.max_row <= 2 and ws.max_column <= 2:
                continue
            rows.extend(parse_format_a_sheet(ws, path.name, sn))

    wb.close()
    return rows


FILE_MONTH_RE = re.compile(
    r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b',
    re.IGNORECASE,
)


def parse_file_primary(filename: str) -> tuple[int, int] | None:
    """Extract the file's primary (year, month) from filename.

    Examples:
        "StudyLink Business Target Apr 2025.xlsx"        → (2025, 4)
        "StudyLink Business Target Mar - Apr 2024.xlsx"  → (2024, 4)
        "StudyLink Business Target Jul - Aug 2023.xlsx"  → (2023, 8)
        "Sub (An + Lợi)_StudyLink Business Target 2025"  → (2025, 12)  # annual file
    """
    # Normalize underscores to spaces so word-boundary matching works
    normalized = filename.replace('_', ' ')
    months = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
              'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
    year_m = re.search(r'(20\d{2})', normalized)
    if not year_m:
        return None
    year = int(year_m.group(1))
    month_matches = FILE_MONTH_RE.findall(normalized)
    if not month_matches:
        return (year, 12)
    last_month = months[month_matches[-1].lower()]
    return (year, last_month)


def aggregate_within_sheet(rows: list[dict]) -> list[dict]:
    """Sum CO pair rows for the same (file, sheet, staff, year, month, type) — collapses
    'Hoàng Yến (w/Vinh)' + 'Hoàng Yến (w/Ngọc Hân)' into one Hoàng Yến row.
    """
    bucket = {}
    for r in rows:
        key = (r['source_file'], r['sheet_name'], r['staff_name_short'],
               r['year'], r['month'], r['target_type'], r['target_unit'])
        if key not in bucket:
            bucket[key] = {
                **r,
                'raw_value': [r['raw_value']],
                'notes_list': [r['notes']] if r['notes'] else [],
            }
        else:
            bucket[key]['target_value'] += r['target_value']
            bucket[key]['raw_value'].append(r['raw_value'])
            if r['notes']:
                bucket[key]['notes_list'].append(r['notes'])
    out = []
    for v in bucket.values():
        v['raw_value'] = ' | '.join(v['raw_value'])
        v['notes'] = ' | '.join(v['notes_list']) if v['notes_list'] else ''
        del v['notes_list']
        out.append(v)
    return out


def dedupe_across_sheets(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """When the same (staff, year, month, target_type) is provided by multiple
    sheets across one or more files, choose the most authoritative source.

    Authority ranking (lowest score wins):
      0. Source file's primary month exactly matches the target month
      1. Source file's primary month is later than the target month (retro corrections)
      2. Source file's primary month is earlier than the target month (forward-looking)

    Within each tier, prefer the closest distance.

    Returns (final_rows, conflicts).  Conflicts lists keys where candidate
    values disagreed numerically — useful for review.
    """
    by_key: dict[tuple, list[dict]] = {}
    for r in rows:
        key = (r['staff_name_short'], r['year'], r['month'],
               r['target_type'], r['target_unit'])
        by_key.setdefault(key, []).append(r)

    final = []
    conflicts = []
    for key, candidates in by_key.items():
        if len(candidates) == 1:
            final.append(candidates[0])
            continue

        target_idx = key[1] * 12 + key[2]

        def score(r):
            fp = parse_file_primary(r['source_file'])
            if fp is None:
                return (3, 0, r['source_file'])
            file_idx = fp[0] * 12 + fp[1]
            dist = file_idx - target_idx
            if dist == 0:
                tier = 0
            elif dist > 0:
                tier = 1
            else:
                tier = 2
            return (tier, abs(dist), r['source_file'])

        candidates.sort(key=score)
        winner = candidates[0]
        final.append(winner)

        # Record conflict if numeric values disagree
        values = {round(c['target_value'], 4) for c in candidates}
        if len(values) > 1:
            conflicts.append({
                'staff_name_short': key[0],
                'year': key[1],
                'month': key[2],
                'target_type': key[3],
                'target_unit': key[4],
                'chosen_source': winner['source_file'],
                'chosen_sheet': winner['sheet_name'],
                'chosen_value': winner['target_value'],
                'all_candidates': '; '.join(
                    f"{c['source_file']}#{c['sheet_name']}={c['target_value']}"
                    for c in candidates
                ),
            })

    return final, conflicts


def main():
    if len(sys.argv) < 2:
        print('Usage: python parse_targets.py <input_folder> [output_csv]')
        sys.exit(1)

    folder = Path(sys.argv[1])
    out_csv = Path(sys.argv[2]) if len(sys.argv) >= 3 else Path('staff_targets.csv')

    if not folder.is_dir():
        print(f'Not a directory: {folder}')
        sys.exit(1)

    files = sorted(folder.glob('*.xlsx'))
    files = [f for f in files if not f.name.startswith('~$')]
    print(f'Found {len(files)} xlsx files in {folder}')

    all_rows = []
    for f in files:
        print(f'  {f.name} ({detect_format(f)})')
        rows = parse_file(f)
        print(f'    -> {len(rows)} target rows extracted')
        all_rows.extend(rows)

    print(f'\nTotal raw rows: {len(all_rows)}')
    within = aggregate_within_sheet(all_rows)
    print(f'After within-sheet CO pair aggregation: {len(within)}')
    aggregated, conflicts = dedupe_across_sheets(within)
    print(f'After cross-sheet dedupe: {len(aggregated)}')
    if conflicts:
        print(f'\n*** {len(conflicts)} VALUE CONFLICTS detected — see conflicts.csv ***')
    else:
        print('\nNo cross-sheet value conflicts.')

    fields = ['source_file', 'sheet_name', 'staff_name_short', 'year', 'month',
              'target_type', 'target_value', 'target_unit', 'raw_value', 'notes']
    with open(out_csv, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in sorted(aggregated, key=lambda x: (x['year'], x['month'], x['target_type'], x['staff_name_short'])):
            writer.writerow({k: r.get(k, '') for k in fields})

    print(f'\nWrote {out_csv}')

    if conflicts:
        conflict_csv = out_csv.with_name(out_csv.stem + '_conflicts.csv')
        cfields = ['staff_name_short', 'year', 'month', 'target_type', 'target_unit',
                   'chosen_source', 'chosen_sheet', 'chosen_value', 'all_candidates']
        with open(conflict_csv, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=cfields)
            writer.writeheader()
            for c in sorted(conflicts, key=lambda x: (x['year'], x['month'], x['staff_name_short'])):
                writer.writerow({k: c.get(k, '') for k in cfields})
        print(f'Wrote {conflict_csv}')

    # Summary
    by_type = {}
    by_yearmonth = {}
    for r in aggregated:
        by_type[r['target_type']] = by_type.get(r['target_type'], 0) + 1
        ym = f"{r['year']}-{r['month']:02d}"
        by_yearmonth[ym] = by_yearmonth.get(ym, 0) + 1

    print('\nRows by target_type:')
    for k, v in sorted(by_type.items()):
        print(f'  {k}: {v}')
    print(f'\nMonths covered: {len(by_yearmonth)}')
    print(f'Earliest: {min(by_yearmonth)}  Latest: {max(by_yearmonth)}')


if __name__ == '__main__':
    main()
