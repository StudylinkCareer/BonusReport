"""
backend/tools/regression_compare.py

Bao cao vs Engine regression comparison tool (v4).

Usage from project root with .venv active:
    python -m backend.tools.regression_compare --staff "Phạm Thị Lợi" --year 2024 --month 1
    python -m backend.tools.regression_compare --staff "Lê Thị Trường An" --year 2024 --month 2 \
        --bao-cao-path "C:/path/to/specific_bao_cao.xlsx"

Three figures per row, mirroring bao cao layout
-----------------------------------------------
  * Net bonus / "BONUS Enrolled"  =  net_payable - priority_bonus
  * Priority bonus                =  priority_bonus
  * Total                         =  net_payable + COALESCE(mgmt_override_amount, 0)

Auto-discovery picks the SHORTEST matching filename — the canonical monthly
bao cao tends to be the unsuffixed one. Excel lock files (~$prefix) are
excluded. Use --bao-cao-path to pick a specific variant explicitly.
"""
import argparse
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from backend.data.connection import get_connection
except ImportError:
    print("ERROR: must be run from the BonusReport project root with .venv active.")
    sys.exit(1)


# ---------------- bao cao discovery / reading ----------------

BAO_CAO_DIR = r"C:\Users\rhod_\Documents\Bonus Engine Support\Bonus Reports"


def _nfc(s: str) -> str:
    return unicodedata.normalize('NFC', s)


_THANG_RE = re.compile(r'tháng[\s_\-]+(\d{1,2})[\s_\-.]+(\d{4})')


def find_bao_cao(staff: str, year: int, month: int) -> str | None:
    """Auto-discover the bao cao path.

    - Tolerant of spaces/underscores in filenames, NFC vs NFD encoding,
      and dot/underscore between month/year.
    - Excel lock files (~$...) are always excluded.
    - When multiple match, picks the SHORTEST filename (canonical bao cao
      is unsuffixed; longer ones are regional/sub-agent splits).
    """
    staff_nfc = _nfc(staff)
    staff_forms = {staff_nfc, staff_nfc.replace(' ', '_')}
    candidates = []
    inspected_xlsx = []

    for root_dir in [BAO_CAO_DIR, '/mnt/project']:
        if not os.path.isdir(root_dir):
            continue
        for dirpath, _, filenames in os.walk(root_dir):
            for f in filenames:
                if not f.lower().endswith('.xlsx'):
                    continue
                if f.startswith('~$'):
                    continue   # Excel lock file

                f_nfc = _nfc(f)
                inspected_xlsx.append((dirpath, f_nfc))

                if not any(form in f_nfc for form in staff_forms):
                    continue

                m = _THANG_RE.search(f_nfc)
                if not m:
                    continue
                f_month, f_year = int(m.group(1)), int(m.group(2))
                if f_month == month and f_year == year:
                    candidates.append(os.path.join(dirpath, f))

    if not candidates:
        print(f"\n[diagnostic] No bao cao matched. Searched in:")
        print(f"  - {BAO_CAO_DIR}")
        print(f"  - /mnt/project")
        if inspected_xlsx:
            print(f"\n[diagnostic] Looked at {len(inspected_xlsx)} .xlsx files. First 10:")
            for dp, f in inspected_xlsx[:10]:
                short_dp = dp if len(dp) <= 70 else '...' + dp[-67:]
                print(f"  {short_dp}\\{f}")
            if len(inspected_xlsx) > 10:
                print(f"  ... and {len(inspected_xlsx) - 10} more")
        else:
            print(f"\n[diagnostic] No .xlsx files found at all under those roots.")
        return None

    # Prefer shortest filename (canonical monthly bao cao is unsuffixed)
    candidates.sort(key=lambda p: (len(os.path.basename(p)), p))
    chosen = candidates[0]

    if len(candidates) > 1:
        print(f"Multiple bao caos found for {staff} {year}-{month:02d}:")
        for c in candidates:
            marker = '  <-- chosen (shortest = canonical)' if c == chosen else ''
            print(f"  {c}{marker}")
        print(f"(Override with --bao-cao-path to pick a different variant.)")
    return chosen


def find_header_row(ws):
    for r in range(1, min(15, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == 'Contract ID':
                return r
    return None


def header_map(ws, hr):
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hr, column=c).value
        if not isinstance(v, str):
            continue
        norm = re.sub(r'\s+', ' ', v.strip())
        if norm.startswith('Note '):
            continue
        h[norm] = c
    return h


def to_int(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(',', '').strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def read_bao_cao(path: str, staff: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    records = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        hr = find_header_row(ws)
        if not hr:
            continue
        cols = header_map(ws, hr)
        if 'Contract ID' not in cols:
            continue

        for r in range(hr + 1, ws.max_row + 1):
            cid = ws.cell(row=r, column=cols['Contract ID']).value
            if not cid:
                continue

            co = ws.cell(row=r, column=cols.get('Case Officer Name', 0)).value if 'Case Officer Name' in cols else None
            couns = ws.cell(row=r, column=cols.get('Counsellor Name', 0)).value if 'Counsellor Name' in cols else None
            if co != staff and couns != staff:
                continue

            cstat = ws.cell(row=r, column=cols.get('Application Report Status', 0)).value if 'Application Report Status' in cols else None
            inst = ws.cell(row=r, column=cols.get('Institution Name', 0)).value if 'Institution Name' in cols else None
            b_enrol = to_int(ws.cell(row=r, column=cols.get('BONUS Enrolled', 0)).value) if 'BONUS Enrolled' in cols else 0
            b_prio = to_int(ws.cell(row=r, column=cols.get('BONUS Priority', 0)).value) if 'BONUS Priority' in cols else 0

            note_e = note_p = None
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(row=hr, column=c).value
                if isinstance(hv, str):
                    hvn = re.sub(r'\s+', ' ', hv.strip())
                    if hvn == 'Note BONUS Enrolled':
                        note_e = ws.cell(row=r, column=c).value
                    if hvn == 'Note BONUS Priority':
                        note_p = ws.cell(row=r, column=c).value

            new_rec = {
                'enrol': b_enrol,
                'prio': b_prio,
                'status': cstat,
                'inst': str(inst) if inst else '',
                'note_e': note_e,
                'note_p': note_p,
                'sheet': sn,
            }
            existing = records.get(cid)
            if existing is None or (existing['enrol'] + existing['prio']) < (b_enrol + b_prio):
                records[cid] = new_rec

    return records


# ---------------- engine reading ----------------

def read_engine_payments(staff_name: str, year: int, month: int) -> dict:
    """
    Pull tx_bonus_payment rows for this staff/year/month.
    Aggregates multiple slot rows per contract by summing.
    Status comes from tx_case.application_status (string column on tx_case).
    """
    sql = """
    SELECT
        c.contract_id,
        s.canonical_name AS staff_name,
        bp.gross_bonus,
        bp.net_payable,
        bp.priority_bonus,
        bp.mgmt_override_amount,
        bp.mgmt_override_reason,
        c.application_status,
        i.canonical_name AS inst_name
    FROM tx_bonus_payment bp
    JOIN ref_staff       s ON s.id = bp.staff_id
    JOIN tx_case         c ON c.id = bp.case_id
    LEFT JOIN ref_institution i ON i.id = c.institution_id
    WHERE s.canonical_name = %(staff_name)s
      AND bp.run_year      = %(year)s
      AND bp.run_month     = %(month)s
    ORDER BY c.contract_id;
    """

    out = {}
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, {'staff_name': staff_name, 'year': year, 'month': month})
            for row in cur:
                cid = row['contract_id']
                net_payable = int(row['net_payable'] or 0)
                priority = int(row['priority_bonus'] or 0)
                override = int(row['mgmt_override_amount'] or 0)
                enrolled = net_payable - priority
                total = net_payable + override

                if cid in out:
                    out[cid]['enrolled'] += enrolled
                    out[cid]['priority'] += priority
                    out[cid]['override'] += override
                    out[cid]['total'] += total
                    if row['mgmt_override_reason']:
                        existing = out[cid].get('override_reason') or ''
                        sep = ' | ' if existing else ''
                        out[cid]['override_reason'] = existing + sep + row['mgmt_override_reason']
                else:
                    out[cid] = {
                        'enrolled': enrolled,
                        'priority': priority,
                        'override': override,
                        'total':    total,
                        'application_status': row['application_status'],
                        'inst':              row['inst_name'],
                        'override_reason':   row['mgmt_override_reason'],
                    }
    return out


# ---------------- comparison ----------------

def write_comparison(out_path: str, staff: str, year: int, month: int,
                     bao_cao: dict, engine: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{staff[:15]} {year}-{month:02d}"

    headers = [
        'Contract ID', 'Institution', 'Bao cao status',
        'Bao cao Enrolled', 'Bao cao Priority', 'Bao cao Total',
        'Engine Enrolled', 'Engine Priority', 'Engine Override', 'Engine Total',
        'Gap (Engine Total - Bao cao Total)',
        'Engine status', 'Verdict',
        'Override reason',
        'Bao cao note (enrol)', 'Bao cao note (prio)',
    ]
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = header_font
        ws.cell(row=1, column=c).fill = header_fill

    all_cids = sorted(set(bao_cao) | set(engine))

    matched = gap_count = bao_only = engine_only = 0
    total_bc = total_en = 0

    for cid in all_cids:
        bc = bao_cao.get(cid)
        en = engine.get(cid)

        bc_enrol = bc['enrol'] if bc else 0
        bc_prio = bc['prio'] if bc else 0
        bc_total = bc_enrol + bc_prio

        en_enrolled = en['enrolled'] if en else 0
        en_priority = en['priority'] if en else 0
        en_override = en['override'] if en else 0
        en_total = en['total'] if en else 0

        gap = en_total - bc_total

        if bc and en:
            if gap == 0:
                verdict = 'MATCH'
                matched += 1
            else:
                verdict = f'GAP {gap:+,}'
                gap_count += 1
        elif bc and not en:
            verdict = 'BAO CAO ONLY (engine missed)'
            bao_only += 1
        elif en and not bc:
            verdict = 'ENGINE ONLY (bao cao zero)'
            engine_only += 1
        else:
            verdict = ''

        total_bc += bc_total
        total_en += en_total

        row = [
            cid,
            (bc['inst'] if bc else (en.get('inst') or '')) if (bc or en) else '',
            bc['status'] if bc else '',
            bc_enrol, bc_prio, bc_total,
            en_enrolled, en_priority, en_override, en_total,
            gap,
            (en.get('application_status') or '') if en else '',
            verdict,
            (en.get('override_reason') or '') if en else '',
            (bc.get('note_e') or '') if bc else '',
            (bc.get('note_p') or '') if bc else '',
        ]
        ws.append(row)

        if verdict.startswith('GAP') or verdict.startswith('BAO CAO') or verdict.startswith('ENGINE ONLY'):
            yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).fill = yellow

    ws.insert_rows(1, 7)
    ws.cell(row=1, column=1, value=f'Regression: {staff} | {year}-{month:02d}').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f'Matched (engine == bao cao):   {matched}')
    ws.cell(row=3, column=1, value=f'Gap (engine != bao cao):       {gap_count}')
    ws.cell(row=4, column=1, value=f'Bao-cao-only (engine missed):  {bao_only}')
    ws.cell(row=5, column=1, value=f'Engine-only (bao cao zero):    {engine_only}')
    ws.cell(row=6, column=1, value=f'Bao cao total: {total_bc:,}   Engine total: {total_en:,}')
    ws.cell(row=7, column=1, value=f'Net gap: {total_en - total_bc:+,} (positive = engine over, negative = engine under)')

    widths = [13, 30, 30, 14, 14, 14, 14, 14, 14, 14, 18, 28, 32, 40, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A9'

    wb.save(out_path)
    print(f"\nWrote regression report: {out_path}")
    print(f"  Matched:            {matched}")
    print(f"  Gap:                {gap_count}")
    print(f"  Bao-cao-only:       {bao_only}")
    print(f"  Engine-only:        {engine_only}")
    print(f"  Total bao cao:      {total_bc:>14,}")
    print(f"  Total engine:       {total_en:>14,}")
    print(f"  Net gap:            {total_en - total_bc:>+14,}")


# ---------------- main ----------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument('--staff', required=True, help='Canonical staff name, e.g. "Phạm Thị Lợi"')
    p.add_argument('--year', type=int, required=True)
    p.add_argument('--month', type=int, required=True)
    p.add_argument('--bao-cao-path', help='Override auto-discovery of bao cao file')
    p.add_argument('--out', help='Output xlsx path; default = regression_<staff>_<year>_<month>.xlsx in cwd')
    args = p.parse_args()

    bc_path = args.bao_cao_path or find_bao_cao(args.staff, args.year, args.month)
    if not bc_path:
        print(f"\nERROR: no bao cao found for {args.staff} {args.year}-{args.month:02d}")
        print(f"Use --bao-cao-path to specify directly.")
        sys.exit(1)
    print(f"Bao cao:    {bc_path}")

    bao_cao = read_bao_cao(bc_path, args.staff)
    print(f"Bao cao cases for {args.staff}: {len(bao_cao)}")

    print(f"Pulling engine payments from DB...")
    engine = read_engine_payments(args.staff, args.year, args.month)
    print(f"Engine cases: {len(engine)}")

    out = args.out or f"regression_{args.staff.replace(' ', '_')}_{args.year}_{args.month:02d}.xlsx"
    write_comparison(out, args.staff, args.year, args.month, bao_cao, engine)


if __name__ == '__main__':
    main()
