"""
Distinct-value extractor for closed-file CRM reports.

Reads every *report_of_closed_file_in_*.xlsx under a given directory
(recursively, so subfolders are scanned too) and produces a single
multi-tab Excel file showing every DISTINCT value found per
column-of-interest, with frequency counts and example source files.

This is a READ-ONLY tool — no DB writes. Its output is a marked-up
spreadsheet that a human reviewer (you / QM) annotates with canonical
mappings, which a separate seeder script then loads into ref tables.

Usage:
    python -m tools.extract_distinct_values \\
        --input-dir "C:/Users/.../closed_file_reports" \\
        --output    "distinct_values_for_review.xlsx"

Or with defaults (project dir → backend/output/distinct_values.xlsx):
    python -m tools.extract_distinct_values

Columns scanned (one Excel tab per column):
    F  Client Type
    G  Country of Study
    H  Refer Source Agent       ← the overloaded one (partner/sub-agent/office)
    I  System Type
    J  Application Report Status
    L  Institution Name
    O  Counsellor Name
    P  Case Officer Name

Each tab columns:
    Distinct Value          The exact string seen in the data
    Frequency               How many times it appears across all files
    Files Seen In           First few filenames containing this value
    Suggested Resolution    BLANK — for human reviewer to fill
    Canonical ID            BLANK — for reviewer to fill
    Notes                   BLANK — for reviewer to fill

Section dividers ("Closed files", "Enrolled" alone in column 1) are
skipped per Q4 from the importer design discussion.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# What we scan
# ---------------------------------------------------------------------------

# Column letter → friendly tab name. Order here is the tab order in the output.
COLUMNS_OF_INTEREST: dict[str, str] = {
    'F': 'Client Type',
    'G': 'Country of Study',
    'H': 'Refer Source Agent',
    'I': 'System Type',
    'J': 'Application Report Status',
    'L': 'Institution Name',
    'O': 'Counsellor Name',
    'P': 'Case Officer Name',
}

# Tab names get sanitized (Excel limits 31 chars, no /\?*[]:).
# Reduce friendly names to safe sheet names.
def _safe_sheet_name(s: str) -> str:
    s = re.sub(r'[\/\\?*\[\]:]', '', s)
    return s[:31]


# ---------------------------------------------------------------------------
# Per-file scanner
# ---------------------------------------------------------------------------

def scan_file(path: Path) -> dict[str, list[str]]:
    """
    Scan one closed_file report. Returns:
        { column_letter : [list of values seen in this file, in order] }

    Skips:
      - Header rows (rows 1 and 2)
      - Section dividers (rows where only column A is populated)
      - Empty cells

    Values are stripped but otherwise preserved as-is — including
    case, whitespace within, and special characters. The whole point
    of this scan is to surface the messy reality.
    """
    result: dict[str, list[str]] = {col: [] for col in COLUMNS_OF_INTEREST}

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [WARN] Could not open {path.name}: {e}", file=sys.stderr)
        return result

    # The Student Contract sheet is canonical, but be defensive.
    sheet_name = 'Student Contract' if 'Student Contract' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # iter_rows in read_only mode is the fast path — yields tuples, not cells.
    # We need the column letters, so we work with index-based access.
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip title and header
        if row_idx <= 2:
            continue

        # Section divider detection: column A is populated, columns B+ are all empty.
        # row[0] is column A. If it's set and the rest of the row is empty, it's a divider.
        if row and row[0] is not None:
            non_a_cells = row[1:]
            if all(cell is None or str(cell).strip() == '' for cell in non_a_cells):
                continue

        # Pull each column of interest.
        for col_letter in COLUMNS_OF_INTEREST:
            col_idx = ord(col_letter) - ord('A')
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            sval = str(val).strip()
            if not sval:
                continue
            result[col_letter].append(sval)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Aggregation across files
# ---------------------------------------------------------------------------

class ValueAggregate:
    """Tracks frequency and source files for one distinct value."""
    __slots__ = ('frequency', 'files')

    def __init__(self) -> None:
        self.frequency: int = 0
        self.files: list[str] = []  # ordered, deduplicated

    def add(self, filename: str) -> None:
        self.frequency += 1
        if filename not in self.files:
            self.files.append(filename)


def aggregate(
    file_results: dict[str, dict[str, list[str]]],
) -> dict[str, dict[str, ValueAggregate]]:
    """
    Combine per-file results into per-column aggregates.

    Input:  { filename : { col_letter : [values...] } }
    Output: { col_letter : { distinct_value : ValueAggregate } }

    Note: filename here is whatever the caller passed — typically the
    relative path from input_dir, which lets reviewers see which
    subfolder a value came from when the input is hierarchical.
    """
    out: dict[str, dict[str, ValueAggregate]] = {
        col: defaultdict(ValueAggregate) for col in COLUMNS_OF_INTEREST
    }
    for filename, per_col in file_results.items():
        for col_letter, values in per_col.items():
            for val in values:
                out[col_letter][val].add(filename)
    return out


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

# Columns we put in each output tab — header row.
OUTPUT_COLUMNS = (
    'Distinct Value',
    'Frequency',
    'Files Seen In',
    'Suggested Resolution',
    'Canonical ID',
    'Notes',
)

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')


def write_output(
    aggregates: dict[str, dict[str, ValueAggregate]],
    summary: dict[str, int],
    output_path: Path,
) -> None:
    """Write the multi-tab review workbook."""
    wb = Workbook()

    # Summary tab first
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    ws_summary['A1'] = 'BonusReport — Distinct-Value Extraction Summary'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A3'] = 'Files scanned'
    ws_summary['B3'] = summary['files_scanned']
    ws_summary['A4'] = 'Files failed to read'
    ws_summary['B4'] = summary['files_failed']
    ws_summary['A6'] = 'Column'
    ws_summary['B6'] = 'Friendly Name'
    ws_summary['C6'] = 'Distinct Values'
    ws_summary['D6'] = 'Total Occurrences'
    for cell in ('A6', 'B6', 'C6', 'D6'):
        ws_summary[cell].font = HEADER_FONT
        ws_summary[cell].fill = HEADER_FILL
    row = 7
    for col_letter, friendly_name in COLUMNS_OF_INTEREST.items():
        agg = aggregates.get(col_letter, {})
        distinct_count = len(agg)
        total_occurrences = sum(v.frequency for v in agg.values())
        ws_summary[f'A{row}'] = col_letter
        ws_summary[f'B{row}'] = friendly_name
        ws_summary[f'C{row}'] = distinct_count
        ws_summary[f'D{row}'] = total_occurrences
        row += 1
    for col_idx, width in enumerate([10, 30, 18, 20], start=1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

    # One tab per column
    for col_letter, friendly_name in COLUMNS_OF_INTEREST.items():
        ws = wb.create_sheet(_safe_sheet_name(friendly_name))
        # Header row
        for col_idx, header in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        # Data rows — sort by frequency desc, then value asc
        col_aggs = aggregates.get(col_letter, {})
        sorted_values = sorted(
            col_aggs.items(),
            key=lambda kv: (-kv[1].frequency, kv[0].lower()),
        )
        for row_idx, (value, agg) in enumerate(sorted_values, start=2):
            ws.cell(row=row_idx, column=1, value=value)
            ws.cell(row=row_idx, column=2, value=agg.frequency)
            # Show first 3 files; abbreviate if more.
            files_shown = agg.files[:3]
            files_str = '; '.join(files_shown)
            if len(agg.files) > 3:
                files_str += f' (+{len(agg.files) - 3} more)'
            ws.cell(row=row_idx, column=3, value=files_str)
            # Reviewer columns left blank.
        # Column widths
        widths = [50, 12, 60, 25, 15, 50]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        # Freeze header row
        ws.freeze_panes = 'A2'

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__.split('\n\n')[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        '--input-dir', type=Path, required=True,
        help='Directory containing *report_of_closed_file_in_*.xlsx files',
    )
    parser.add_argument(
        '--output', type=Path,
        default=Path('output/distinct_values_for_review.xlsx'),
        help='Output workbook path (default: output/distinct_values_for_review.xlsx)',
    )
    parser.add_argument(
        '--pattern', type=str,
        default='*report_of_closed_file_in_*.xlsx',
        help='Glob pattern for input files',
    )
    args = parser.parse_args()

    if not args.input_dir.is_dir():
        print(f"[FAIL] Input directory not found: {args.input_dir}",
              file=sys.stderr)
        return 1

    files = sorted(args.input_dir.rglob(args.pattern))
    if not files:
        print(f"[FAIL] No files matched {args.pattern!r} in {args.input_dir}",
              file=sys.stderr)
        return 1

    print(f"Scanning {len(files)} file(s) under {args.input_dir}...")
    file_results: dict[str, dict[str, list[str]]] = {}
    failed = 0
    for path in files:
        per_col = scan_file(path)
        if not any(per_col.values()):
            failed += 1
        # Use relative path so the reviewer can see which subfolder
        # a value came from. Falls back to just the name if for some
        # reason the relative computation fails (shouldn't happen).
        try:
            display = str(path.relative_to(args.input_dir))
        except ValueError:
            display = path.name
        file_results[display] = per_col

    aggregates = aggregate(file_results)

    # Ensure output dir exists
    args.output.parent.mkdir(parents=True, exist_ok=True)
    summary = {
        'files_scanned': len(files),
        'files_failed': failed,
    }
    write_output(aggregates, summary, args.output)

    # Console summary
    print()
    print(f"  Files scanned: {len(files)}")
    if failed:
        print(f"  Files failed:  {failed}")
    print(f"  Output: {args.output}")
    print()
    print("  Distinct value counts per column:")
    for col_letter, friendly_name in COLUMNS_OF_INTEREST.items():
        n = len(aggregates.get(col_letter, {}))
        print(f"    {col_letter}  {friendly_name:30s} {n:>4} distinct")

    return 0


if __name__ == '__main__':
    sys.exit(main())
